"""Import master data from XLSX. Clears existing master tables first."""
from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.extensions import db
from app.models import (
    ApprovalDecision,
    ApprovalFlow,
    ApprovalParticipant,
    AuditLog,
    Comment,
    Department,
    Document,
    DocumentVersion,
    DocumentPermission,
    Notification,
    Position,
    Space,
    User,
)
import json
import uuid


def _create_tiptap_doc(content: list) -> dict:
    return {"type": "doc", "content": content}

def _create_heading(text: str, level: int = 1) -> dict:
    return {
        "type": "heading",
        "attrs": {"level": level},
        "content": [{"type": "text", "text": text}]
    }

def _create_paragraph(text: str = "") -> dict:
    item = {"type": "paragraph"}
    if text:
        item["content"] = [{"type": "text", "text": text}]
    return item

def _create_bullet_list(items: list[str]) -> dict:
    return {
        "type": "bulletList",
        "content": [
            {
                "type": "listItem",
                "content": [{"type": "paragraph", "content": [{"type": "text", "text": i}]}]
            } for i in items
        ]
    }

def init_standard_templates(admin_user_id: int | None = None):
    templates_data = [
        {
            "title": "Employment Contract",
            "content": _create_tiptap_doc([
                _create_heading("EMPLOYMENT AGREEMENT"),
                _create_paragraph("This Agreement is made on [DATE] between:"),
                _create_paragraph("PARTY A (EMPLOYER): [Organization Name]"),
                _create_paragraph("PARTY B (EMPLOYEE): [Full Name]"),
                _create_heading("1. Position and Responsibilities", 2),
                _create_paragraph("The Employee shall perform the following duties:"),
                _create_bullet_list(["Manage organizational workflow", "Report to department head", "Ensure quality compliance"]),
                _create_heading("2. Term and Compensation", 2),
                _create_paragraph("Salary: [Amount] per month. Effective Date: [Start Date].")
            ])
        },
        {
            "title": "Meeting Minutes",
            "content": _create_tiptap_doc([
                _create_heading("MEETING MINUTES"),
                _create_paragraph("DATE: [Enter Date] | TIME: [Enter Time]"),
                _create_paragraph("TOPIC: [Project/Department Meeting]"),
                _create_heading("Attendees", 2),
                _create_bullet_list(["[Name 1]", "[Name 2]", "[Name 3]"]),
                _create_heading("Agenda & Discussion", 2),
                _create_paragraph("[Details of discussions...]"),
                _create_heading("Action Items", 2),
                _create_bullet_list(["Task A - Assigned to [Name]", "Task B - Due [Date]"])
            ])
        },
        {
            "title": "Technical Specification",
            "content": _create_tiptap_doc([
                _create_heading("TECHNICAL SPECIFICATION"),
                _create_paragraph("VERSION: 1.0 | STATUS: Draft"),
                _create_heading("1. Overview", 2),
                _create_paragraph("Provide a high-level summary of the system or feature."),
                _create_heading("2. Requirements", 2),
                _create_bullet_list(["Requirement 1", "Requirement 2", "Requirement 3"]),
                _create_heading("3. Design & Architecture", 2),
                _create_paragraph("Describe components, data flow, and technologies used.")
            ])
        },
        {
            "title": "Weekly Report",
            "content": _create_tiptap_doc([
                _create_heading("WEEKLY PROGRESS REPORT"),
                _create_paragraph("PERIOD: [Start Date] to [End Date]"),
                _create_heading("Done This Week", 2),
                _create_bullet_list(["Completed Task A", "Delivered Feature B"]),
                _create_heading("Planned for Next Week", 2),
                _create_bullet_list(["Start Project C", "Review Milestone D"]),
                _create_heading("Risks & Blockers", 2),
                _create_paragraph("None.")
            ])
        },
        {
            "title": "Expense Report",
            "content": _create_tiptap_doc([
                _create_heading("EMPLOYEE EXPENSE REIMBURSEMENT"),
                _create_paragraph("SUBMITTED BY: [Employee name]"),
                _create_paragraph("DEPARTMENT: [Department]"),
                _create_heading("Expense Summary", 2),
                _create_paragraph("DATE | DESCRIPTION | CATEGORY | AMOUNT"),
                _create_paragraph("-------------------------------------------"),
                _create_paragraph("[Date] | [Travel/Lunch] | [Category] | [0.00]"),
                _create_paragraph("TOTAL: 0.00")
            ])
        }
    ]
    
    count = 0
    for tmpl in templates_data:
        doc = Document(
            owner_id=admin_user_id,
            title=tmpl["title"],
            status="approved",
            doc_number=f"TMPL-{uuid.uuid4().hex[:6].upper()}",
            is_template=True,
            is_public=True
        )
        db.session.add(doc)
        db.session.flush()
        
        ver = DocumentVersion(
            document_id=doc.id,
            version_no=1,
            content_json=json.dumps(tmpl["content"]),
            created_by_id=admin_user_id
        )
        db.session.add(ver)
        db.session.flush()
        doc.current_version_id = ver.id
        count += 1
    return count

def _norm_key(s: str) -> str:
    return re.sub(r"\s+", "", str(s).strip().lower()) if s is not None else ""


def _match_header(header_row: tuple[Any, ...]) -> dict[str, int]:
    """Map canonical field names to column indices."""
    mapping: dict[str, int] = {}
    aliases = [
        (
            "employee_no",
            [
                "员工编号",
                "empno",
                "employeeno",
                "employeeid",
            ],
        ),
        ("last_name", ["姓(LastName)", "lastname", "surname", "姓 "]),
        ("first_name", ["名(FirstName)", "firstname", "名 "]),
        ("patronymic", ["父称", "middlename"]),
        ("birth_date", ["出生日期", "birthdate"]),
        ("gender", ["性别", "gender"]),
        ("login_name", ["登录名", "login", "loginname", "用户名"]),
        (
            "department_code",
            [
                "部门编号",
                "deptcode",
                "departmentcode",
            ],
        ),
        (
            "department_name",
            ["部门名称", "departmentname"],
        ),
        (
            "department_name_en",
            ["部门英文名", "departmentnameen"],
        ),
        (
            "position_short",
            [
                "职务简称",
                "position",
            ],
        ),
        (
            "position_full",
            ["职务全称", "positionname"],
        ),
        (
            "position_full_en",
            ["职务英文全称", "positionfullnameen"],
        ),
        (
            "manager_employee_no",
            ["直属上级员工编号", "manager"],
        ),
    ]
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        raw = str(cell).strip()
        nk = _norm_key(raw)
        # 1. 优先尝试精确匹配
        found = False
        for canonical, labels in aliases:
            for lb in labels:
                if _norm_key(lb) == nk:
                    mapping[canonical] = idx
                    found = True
                    break
            if found: break
        
        # 2. 如果没找到精确的，再尝试包含匹配（但要排除一些太短的容易误判的情况）
        if not found:
            for canonical, labels in aliases:
                for lb in labels:
                    lb_norm = _norm_key(lb)
                    if lb_norm and lb_norm in nk:
                        mapping[canonical] = idx
                        found = True
                        break
                if found: break
    return mapping


def _parse_date(val: Any):
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d", "%m/%d/%y", "%m/%d/%Y", "%d/%m/%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def import_master_data_xlsx(file_bytes: bytes, overwrite: bool = True, table_type: str = "all") -> dict[str, Any]:
    """Replace or append master data with content from workbook."""
    print(f"[IMPORT] Starting import process. Mode: {overwrite}, Type: {table_type}")
    
    try:
        wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
        print(f"[IMPORT] Workbook loaded. Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"[IMPORT] FAILED to load workbook: {e}")
        raise

    dept_rows: list[tuple[str, str]] = []
    pos_rows: list[tuple[str, str]] = []
    emp_rows: list[dict[str, Any]] = []
    mgr_rows: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        print(f"[IMPORT] Processing sheet: {sheet_name}")
        sn = _norm_key(sheet_name)
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]

        # 💡 根据 table_type 决定是否处理当前 Sheet
        is_dept = ("部门" in sheet_name or sn == "departments" or sn == "dept")
        is_pos = ("职务" in sheet_name or "position" in sn)
        is_emp = ("员工" in sheet_name or "employ" in sn or sn == "staff")
        is_mgr = ("管理" in sheet_name or "manager" in sn)

        if (table_type == "all" or table_type == "departments") and is_dept:
            keys = _match_header(header)
            if "department_code" in keys and "department_name" in keys:
                for r in rows[1:]:
                    code = str(r[keys["department_code"]]).strip() if r[keys["department_code"]] is not None else ""
                    name = str(r[keys["department_name"]] or code).strip()
                    name_en = str(r[keys.get("department_name_en")] or name).strip() if keys.get("department_name_en") is not None else None
                    if code:
                        dept_rows.append((code, name, name_en))
            elif len(header) >= 2:
                for r in rows[1:]:
                    if r and r[0] is not None:
                        dept_rows.append((str(r[0]).strip(), str(r[1] or r[0]).strip(), None))

        elif (table_type == "all" or table_type == "positions") and is_pos:
            keys = _match_header(header)
            if "position_short" in keys and "position_full" in keys:
                for r in rows[1:]:
                    if not r: continue
                    ps = r[keys["position_short"]]
                    if ps is None: continue
                    pf = r[keys["position_full"]] or ps
                    pf_en = r[keys.get("position_full_en")] or pf if keys.get("position_full_en") is not None else None
                    pos_rows.append((str(ps).strip(), str(pf).strip(), str(pf_en).strip() if pf_en else None))
            elif len(header) >= 2:
                for r in rows[1:]:
                    if r and r[0] is not None:
                        pos_rows.append((str(r[0]).strip(), str(r[1] or r[0]).strip(), None))

        elif (table_type == "all" or table_type == "employees") and (is_emp or is_mgr):
            keys = _match_header(header)
            print(f"[IMPORT]   Employees/Managers Match: {keys}")
            if "login_name" in keys:
                for r in rows[1:]:
                    if not r or r[keys["login_name"]] is None:
                        continue
                    rowd = {k: (r[ci] if ci < len(r) else None) for k, ci in keys.items()}
                    rowd["_is_manager"] = is_mgr or rowd.get("is_manager", False)
                    emp_rows.append(rowd)
            else:
                # 💡 如果还是识别失败，尝试针对特定格式的 Fallback
                if len(header) >= 4:
                    print("[IMPORT]   Applying Fallback parser for Employees due to header mismatch...")
                    # 假设 0为login, 1为emp_no, 4为dept, 5为pos
                    for r in rows[1:]:
                        if not r or not r[0]: continue
                        emp_rows.append({
                            "login_name": str(r[0]).strip(),
                            "employee_no": str(r[1] or r[0]).strip(),
                            "last_name": str(r[2] or "-").strip(),
                            "first_name": str(r[3] or "-").strip(),
                            "department_code": str(r[4]).strip() if len(r)>4 and r[4] else None,
                            "position_short": str(r[5]).strip() if len(r)>5 and r[5] else None,
                            "manager_employee_no": str(r[6]).strip() if len(r)>6 and r[6] else None,
                            "gender": str(r[7]).strip() if len(r)>7 and r[7] else None,
                            "_is_manager": is_mgr
                        })

    print(f"[IMPORT] Parsing complete. Depts: {len(dept_rows)}, Pos: {len(pos_rows)}, Employees/Managers: {len(emp_rows)}+{len(mgr_rows)}")

    errors: list[str] = []

    def flush_all():
        """Clear business + master data so user FKs can be recreated. PROTECT ADMIN."""
        from sqlalchemy import delete, select
        print("[IMPORT] FLUSH_ALL: Clearing all business and master data tables (Protecting Admin)...")
        db.session.execute(delete(ApprovalDecision))
        db.session.execute(delete(ApprovalParticipant))
        db.session.execute(delete(ApprovalFlow))
        db.session.execute(delete(AuditLog))
        db.session.execute(delete(Notification))
        db.session.execute(delete(Comment))
        db.session.execute(delete(DocumentPermission))
        db.session.execute(delete(Space))
        for doc in db.session.scalars(select(Document)).all():
            db.session.delete(doc)
        
        # 保护 admin 用户，删除其他所有用户
        db.session.execute(delete(User).where(User.is_super_admin == False))
        
        db.session.execute(delete(Position))
        db.session.execute(delete(Department))
        
        # 确保 admin 用户依然是 active 且为 manager
        admin = db.session.query(User).filter_by(is_super_admin=True).first()
        if admin:
            admin.is_manager = True
            admin.registration_status = 'active'
        
        db.session.flush()

    if overwrite:
        flush_all()
    else:
        print("[IMPORT] APPEND MODE: Skipping flush_all.")

    code_to_dept: dict[str, Department] = {}
    
    # Pre-map existing departments if in append mode
    if not overwrite:
        from sqlalchemy import select
        for d in db.session.scalars(select(Department)).all():
            code_to_dept[d.code] = d

    print("[IMPORT] Synchronizing Departments...")
    for code, name, name_en in dept_rows:
        if not code:
            continue
        if code in code_to_dept:
            d = code_to_dept[code]
            d.name = name
            d.name_en = name_en
        else:
            d = Department(code=code, name=name, name_en=name_en)
            db.session.add(d)
            db.session.flush()
            code_to_dept[code] = d

    print("[IMPORT] Synchronizing Positions...")
    seen_pos: dict[str, Position] = {}
    if not overwrite:
        from sqlalchemy import select
        for p in db.session.scalars(select(Position)).all():
            seen_pos[p.short_name] = p

    for ps, pf, pf_en in pos_rows:
        if ps in seen_pos:
            p = seen_pos[ps]
            p.full_name = pf
            p.full_name_en = pf_en
        else:
            p = Position(short_name=ps, full_name=pf, full_name_en=pf_en)
            db.session.add(p)
            db.session.flush()
            seen_pos[ps] = p

    print("[IMPORT] Synchronizing Users...")
    def upsert_user(row: dict[str, Any]) -> None:
        login = str(row.get("login_name") or "").strip()
        if not login:
            return
        
        # ALWAYS check if user exists to prevent unique constraint failures (especially for admin)
        u = User.query.filter_by(login_name=login).first()

        emp_no = str(row.get("employee_no") or login).strip()
        dept_code = row.get("department_code")
        dept_id = None
        if dept_code is not None:
            dc = str(dept_code).strip()
            if dc in code_to_dept:
                dept_id = code_to_dept[dc].id
        
        if u:
            # Update existing user
            u.employee_no = emp_no
            u.last_name = str(row.get("last_name") or u.last_name).strip()
            u.first_name = str(row.get("first_name") or u.first_name).strip()
            u.patronymic = str(row.get("patronymic") or u.patronymic).strip()
            new_birth = _parse_date(row.get("birth_date"))
            if new_birth: u.birth_date = new_birth
            u.gender = str(row.get("gender") or u.gender).strip()
            u.department_id = dept_id or u.department_id
            u.position_short = str(row.get("position_short") or u.position_short).strip()
            u.is_manager = bool(row.get("_is_manager")) or u.is_manager or u.is_super_admin
            u.registration_status = "active" # Force active on import
        else:
            # Create new user
            u = User(
                employee_no=emp_no,
                last_name=str(row.get("last_name") or "-").strip(),
                first_name=str(row.get("first_name") or "-").strip(),
                patronymic=str(row.get("patronymic") or "").strip(),
                birth_date=_parse_date(row.get("birth_date")),
                gender=str(row.get("gender") or "").strip(),
                login_name=login,
                department_id=dept_id,
                position_short=str(row.get("position_short") or "").strip(),
                manager_employee_no=str(row.get("manager_employee_no") or "").strip(),
                is_manager=bool(row.get("_is_manager")) or login == "admin",
                registration_status="active",
            )
            u.set_password("123456")
            db.session.add(u)

    merged: dict[str, dict[str, Any]] = {}
    for row in emp_rows + mgr_rows:
        login = str(row.get("login_name") or "").strip()
        if not login:
            continue
        if login not in merged:
            merged[login] = row
        else:
            merged[login]["_is_manager"] = merged[login].get("_is_manager") or row.get(
                "_is_manager"
            )

    for idx, row in enumerate(merged.values()):
        try:
            upsert_user(row)
            if idx % 10 == 0: print(f"[IMPORT]   Processed {idx} users...")
        except Exception as exc:  # noqa: BLE001
            print(f"[IMPORT]   ERROR on user {row.get('login_name')}: {exc}")
            errors.append(f"{row.get('login_name')}: {exc}")

    db.session.flush()
    print("[IMPORT] Database flushed.")

    # Initialize standard templates after users are created
    admin_user = db.session.query(User).filter_by(is_super_admin=True).first() or db.session.query(User).first()
    admin_id = admin_user.id if admin_user else None
    
    tmpl_count = 0
    if overwrite:
        print("[IMPORT] Initializing standard templates...")
        tmpl_count = init_standard_templates(admin_id)
        print(f"[IMPORT] {tmpl_count} templates created.")

    print(f"[IMPORT] Done. Departments: {len(dept_rows)}, Users: {len(merged)}, Errors: {len(errors)}")

    return {
        "departments": len(dept_rows),
        "positions": len(pos_rows),
        "users": len(merged),
        "templates": tmpl_count,
        "errors": errors,
    }
