"""Admin endpoints for master data import (requires manager authentication)."""
from flask import Blueprint, jsonify, request
from flask_jwt_extended import verify_jwt_in_request

from app.extensions import db
from app.models import User
from app.services.import_xlsx import import_master_data_xlsx
from app.utils.auth import current_user

bp = Blueprint("admin", __name__)


def _has_managers() -> bool:
    return db.session.query(User.id).filter(User.is_manager == True).first() is not None


@bp.get("/status")
def admin_status():
    """Whether DB has managers (for UI hints)."""
    has_users = _has_managers()
    sample_managers = []
    if has_users:
        sample_managers = [
            u.login_name
            for u in db.session.query(User.login_name)
            .filter(User.is_manager == True)
            .limit(3)
            .all()
        ]
    return jsonify(
        {
            "has_users": has_users,
            "sample_managers": sample_managers,
        }
    )



@bp.post("/master-data/import")
def admin_import_master_data():
    """Import XLSX (clears or appends documents and master data)."""
    # 💡 增加：请求诊断日志
    print(f"[DEBUG] Import request received. Content-Type: {request.content_type}")
    print(f"[DEBUG] Files in request: {list(request.files.keys())}")
    
    # Check if there are any managers in the database
    has_managers = _has_managers()
    
    if has_managers:
        try:
            verify_jwt_in_request()
            user = current_user()
            # 💡 优化：仅允许超级管理员（login_name 为 admin）进行 Excel 全量导入
            if not user or not user.is_super_admin:
                return jsonify({
                    "error": "Access denied. Only the super admin can import master data via Excel."
                }), 403
        except Exception:
            return jsonify({
                "error": "Authorization required. Please sign in as a manager to import master data."
            }), 403

    if "file" not in request.files:
        print("[DEBUG] 400 ERROR: 'file' key missing in request.files")
        return jsonify({"error": "file required"}), 400
    
    f = request.files["file"]
    raw = f.read()
    if not raw:
        print("[DEBUG] 400 ERROR: File is empty")
        return jsonify({"error": "empty file"}), 400

    # Get overwrite and table_type flags
    overwrite = request.form.get("overwrite") != "false"
    table_type = request.form.get("table_type", "all") # 'all', 'departments', 'positions', 'employees'
    print(f"[DEBUG] Overwrite flag: {overwrite}, Table Type: {table_type}")

    try:
        stats = import_master_data_xlsx(raw, overwrite=overwrite, table_type=table_type)
        db.session.commit()
    except Exception as exc:  # noqa: BLE001
        db.session.rollback()
        import traceback
        error_trace = traceback.format_exc()
        print(f"[DEBUG] 400 ERROR (Import Failed): {exc}\n{error_trace}")
        return jsonify({"error": f"Import failed: {str(exc)}"}), 400

    try:
        logins = [
            row[0]
            for row in db.session.query(User.login_name)
            .order_by(User.login_name)
            .limit(15)
            .all()
        ]
        stats["sample_login_names"] = logins
    except Exception:
        stats["sample_login_names"] = []

    return jsonify(stats), 200


from app.models.workflow import AuditLog
from flask_jwt_extended import jwt_required
from sqlalchemy import or_

@bp.get("/audit-logs")
@jwt_required()
def admin_list_audit_logs():
    """Get audit logs (System Admin only)"""
    user = current_user()
    if not user or not user.is_super_admin:
        return jsonify({"error": "Forbidden"}), 403

    from datetime import datetime, timedelta
    cutoff = datetime.utcnow() - timedelta(days=10)
    query = AuditLog.query.filter(or_(AuditLog.created_at >= cutoff, AuditLog.is_starred == True))

    # Filters
    doc_number = request.args.get("doc_number")
    action = request.args.get("action")
    user_id = request.args.get("user_id")

    if doc_number:
        from app.models.document import Document
        query = query.join(Document, AuditLog.document_id == Document.id).filter(Document.doc_number == doc_number)
    if action:
        query = query.filter(AuditLog.action == action)
    if user_id:
        query = query.filter(AuditLog.user_id == user_id)

    page = int(request.args.get("page", 1))
    size = int(request.args.get("size", 20))
    
    total = query.count()
    logs = query.order_by(AuditLog.created_at.desc()).offset((page - 1) * size).limit(size).all()

    items = []
    for lg in logs:
        items.append({
            "id": lg.id,
            "document_id": lg.document_id,
            "doc_number": lg.document.doc_number if lg.document else None,
            "document_title": lg.document.title if lg.document else None,
            "user_id": lg.user_id,
            "user_login": lg.user.login_name if lg.user else None,
            "action": lg.action,
            "summary": lg.summary,
            "ip_address": lg.ip_address,
            "is_starred": lg.is_starred,
            "created_at": lg.created_at.isoformat() + "Z" if lg.created_at else None
        })

    return jsonify({"total": total, "items": items}), 200


@bp.post("/audit-logs/<int:log_id>/toggle-star")
@jwt_required()
def admin_toggle_audit_star(log_id: int):
    """Star or unstar an audit log."""
    user = current_user()
    if not user or not user.is_super_admin:
        return jsonify({"error": "Forbidden"}), 403
    
    log = db.session.get(AuditLog, log_id)
    if not log:
        return jsonify({"error": "Not found"}), 404
    
    from datetime import datetime
    if log.is_starred:
        # Unstarring
        log.is_starred = False
        log.unstarred_at = datetime.utcnow()
    else:
        # Starring
        log.is_starred = True
        log.unstarred_at = None
    
    db.session.commit()
    return jsonify({"is_starred": log.is_starred}), 200


@bp.get("/master-data/template")
def admin_get_import_template():
    """Generate a sample XLSX template for the user."""
    from openpyxl import Workbook
    from io import BytesIO
    from flask import send_file

    wb = Workbook()
    
    # Sheet 1: Departments
    ws1 = wb.active
    ws1.title = "Departments"
    ws1.append(["部门编号 (Dept Code)", "部门名称 (Dept Name)", "部门英文名 (Dept Name EN)"])
    ws1.append(["D001", "研发部", "Research & Development"])
    ws1.append(["D002", "人力资源部", "Human Resources"])

    # Sheet 2: Positions
    ws2 = wb.create_sheet("Positions")
    ws2.append(["职务简称 (Short Name)", "职务全称 (Full Name)", "职务英文全称 (Full Name EN)"])
    ws2.append(["DEV", "开发工程师", "Software Developer"])
    ws2.append(["MGR", "经理", "Manager"])

    # Sheet 3: Employees
    ws3 = wb.create_sheet("Employees")
    ws3.append([
        "登录名 (Login Name)", "员工编号 (Emp No)", "姓 (Last Name)", "名 (First Name)", 
        "部门编号 (Dept Code)", "职务简称 (Position)", "直属上级员工编号 (Manager No)", "性别 (Gender)"
    ])
    ws3.append(["user1", "E001", "张", "三", "D001", "DEV", "AD001", "男"])
    ws3.append(["mgr1", "E002", "李", "四", "D001", "MGR", "AD001", "女"])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="edms_import_template.xlsx"
    )
