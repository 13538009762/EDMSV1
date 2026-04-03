import openpyxl
from io import BytesIO
from app import create_app
from app.extensions import db
from app.models import User, Department
from app.services.import_xlsx import import_master_data_xlsx

def create_sample_xlsx():
    wb = openpyxl.Workbook()
    # Sheet 1: Departments
    ws1 = wb.active
    ws1.title = "Departments"
    ws1.append(["Code", "Name"])
    ws1.append(["D1", "Sales"])
    ws1.append(["D2", "Engineering"])
    
    # Sheet 2: Managers
    ws2 = wb.create_sheet("Managers")
    ws2.append(["Login", "EmpNo", "Last Name", "First Name", "DeptCode"])
    ws2.append(["mgr1", "M1", "Smith", "John", "D1"])
    
    # Sheet 3: Employees
    ws3 = wb.create_sheet("Employees")
    ws3.append(["Login", "EmpNo", "Last Name", "First Name", "DeptCode"])
    ws3.append(["emp1", "E1", "Doe", "Jane", "D2"])
    
    out = BytesIO()
    wb.save(out)
    return out.getvalue()

def create_invalid_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Departments"
    ws.append(["Code", "Name"])
    ws.append(["D1", "Duplicate Code"])
    ws.append(["D1", "This should cause an error if uniquely constrained"])
    
    out = BytesIO()
    wb.save(out)
    return out.getvalue()

app = create_app()
with app.app_context():
    print("Testing valid import...")
    raw = create_sample_xlsx()
    # In the actual API, we wrap this in a transaction
    with db.session.begin():
        stats = import_master_data_xlsx(raw)
        print(f"Import stats: {stats}")
    
    user_count = db.session.query(User).count()
    print(f"Users after valid import: {user_count}")
    
    print("\nTesting invalid import (duplicate dept code)...")
    raw_invalid = create_invalid_xlsx()
    try:
        with db.session.begin():
            import_master_data_xlsx(raw_invalid)
        print("Error: Invalid import should have failed!")
    except Exception as e:
        print(f"Successfully caught expected error: {e}")
        # db.session.rollback() is automatic with `with db.session.begin()`
    
    user_count_after = db.session.query(User).count()
    print(f"Users after failed import: {user_count_after}")
    if user_count_after == user_count:
        print("Success: Database was NOT cleared after failure.")
    else:
        print("Failure: Database WAS cleared after failure.")
