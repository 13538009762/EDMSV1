import openpyxl
from openpyxl.styles import Font, PatternFill

def generate_test_xlsx(output_path):
    wb = openpyxl.Workbook()
    
    # 1. Departments Sheet
    ws_dept = wb.active
    ws_dept.title = "Departments"
    headers_dept = ["部门编号 (Dept Code)", "部门名称 (Dept Name)", "部门英文名 (Dept Name EN)"]
    ws_dept.append(headers_dept)
    
    depts = [
        ("HO", "总部", "Head Office"),
        ("TECH", "技术部", "Technology Department"),
        ("HR", "人力资源部", "Human Resources"),
        ("FIN", "财务部", "Finance Department"),
        ("SALES", "销售部", "Sales Department"),
    ]
    for d in depts:
        ws_dept.append(d)
        
    # 2. Positions Sheet
    ws_pos = wb.create_sheet("Positions")
    headers_pos = ["职务简称 (Short Name)", "职务全称 (Full Name)", "职务英文全称 (Full Name EN)"]
    ws_pos.append(headers_pos)
    
    positions = [
        ("CEO", "首席执行官", "Chief Executive Officer"),
        ("CTO", "首席技术官", "Chief Technology Officer"),
        ("DIR", "总监", "Director"),
        ("MGR", "经理", "Manager"),
        ("STF", "员工", "Staff"),
    ]
    for p in positions:
        ws_pos.append(p)
        
    # 3. Employees Sheet
    ws_emp = wb.create_sheet("Employees")
    headers_emp = [
        "登录名 (Login Name)", "员工编号 (Emp No)", "姓 (Last Name)", "名 (First Name)", 
        "部门编号 (Dept Code)", "职务简称 (Position)", "直属上级员工编号 (Manager No)", "性别 (Gender)"
    ]
    ws_emp.append(headers_emp)
    
    employees = [
        ("admin", "EMP001", "系统", "管理员", "HO", "CEO", "", "Male"),
        ("tech_mgr", "EMP002", "张", "技术", "TECH", "DIR", "EMP001", "Male"),
        ("dev_01", "EMP003", "李", "开发", "TECH", "STF", "EMP002", "Male"),
        ("hr_mgr", "EMP004", "王", "人事", "HR", "MGR", "EMP001", "Female"),
        ("sales_01", "EMP005", "赵", "销售", "SALES", "STF", "EMP001", "Female"),
    ]
    for e in employees:
        ws_emp.append(e)
        
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        # Adjust column width
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

    wb.save(output_path)
    print(f"Generated {output_path}")

if __name__ == "__main__":
    import os
    target = os.path.join(os.getcwd(), "test_personnel.xlsx")
    generate_test_xlsx(target)
